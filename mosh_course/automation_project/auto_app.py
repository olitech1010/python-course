import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filenname):
    wb = xl.load_workbook(filenname)
    sheet = wb['Sheet1']
    cell = sheet.cell(1,1)
    new_price_head = sheet.cell(1,4)
    new_price_head.value = 'New Price'

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        new_price = cell.value*3
        new_price_cell = sheet.cell(row, 4)
        new_price_cell.value = new_price

    values = Reference(sheet,
            min_row = 2,
            max_row = sheet.max_row,
            min_col = 4,
            max_col = 4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filenname)
